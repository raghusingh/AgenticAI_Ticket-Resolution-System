from pathlib import Path
from typing import Dict, List
import json
import csv
import re
from io import StringIO

from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook


class SharePointLocalIngestor:
    def extract(self, source: Dict) -> List[str]:
        folder_path = source.get("source_url") or source.get("folder_path")

        if not folder_path:
            raise ValueError("Local SharePoint folder path is missing")

        base_path = Path(folder_path)

        if not base_path.exists():
            raise ValueError(f"Folder not found: {folder_path}")

        docs: List[str] = []

        for file_path in base_path.rglob("*"):
            if not file_path.is_file():
                continue

            suffix = file_path.suffix.lower()

            try:
                if suffix == ".xlsx":
                    docs.extend(self._read_xlsx_as_ticket_docs(file_path))
                    continue

                raw_text = self._read_file(file_path, suffix)

                if not raw_text.strip():
                    continue

                fields = self._extract_fields(raw_text)

                content = f"""
Issue Key: {fields.get("issue_key", "")}
Type: {fields.get("issue_type", "")}
Summary: {fields.get("summary", "")}
Status: {fields.get("status", "")}
Priority: {fields.get("priority", "")}
Created: {fields.get("created", "")}
Updated: {fields.get("updated", "")}

Description:
{raw_text}
""".strip()

                docs.append(content)

            except Exception as ex:
                print(f"Skipping file {file_path}: {ex}")

        return docs

    # ---------------- GENERIC FILE READERS ---------------- #

    def _read_file(self, file_path: Path, suffix: str) -> str:
        if suffix in [".txt", ".log", ".md"]:
            return file_path.read_text(encoding="utf-8", errors="ignore")

        if suffix == ".json":
            raw = file_path.read_text(encoding="utf-8", errors="ignore")
            try:
                obj = json.loads(raw)
                return json.dumps(obj, indent=2, ensure_ascii=False)
            except Exception:
                return raw

        if suffix == ".csv":
            raw = file_path.read_text(encoding="utf-8", errors="ignore")
            return self._csv_to_text(raw)

        if suffix == ".pdf":
            return self._read_pdf(file_path)

        if suffix == ".docx":
            return self._read_docx(file_path)

        return ""

    def _csv_to_text(self, raw_csv: str) -> str:
        reader = csv.reader(StringIO(raw_csv))
        rows = list(reader)

        lines = []
        for row in rows:
            cleaned = [str(cell).strip() for cell in row if cell]
            if cleaned:
                lines.append(" | ".join(cleaned))

        return "\n".join(lines)

    def _read_pdf(self, file_path: Path) -> str:
        reader = PdfReader(str(file_path))
        pages = []

        for page in reader.pages:
            text = page.extract_text() or ""
            if text.strip():
                pages.append(text.strip())

        return "\n\n".join(pages)

    def _read_docx(self, file_path: Path) -> str:
        doc = Document(str(file_path))
        return "\n".join([p.text for p in doc.paragraphs if p.text.strip()])

    # ---------------- XLSX TICKET EXTRACTION ---------------- #

    def _read_xlsx_as_ticket_docs(self, file_path: Path) -> List[str]:
        wb = load_workbook(filename=str(file_path), data_only=True)
        docs: List[str] = []

        for ws in wb.worksheets:
            # Only process ticket-like sheets
            headers = [cell.value for cell in ws[1]]
            header_map = {str(h).strip(): idx for idx, h in enumerate(headers) if h is not None}

            # Detect enterprise incident sheet
            if "Incident_Number" not in header_map:
                continue

            for row in ws.iter_rows(min_row=2, values_only=True):
                incident_number = self._get_cell(row, header_map, "Incident_Number")
                if not incident_number:
                    continue

                issue_type = self._get_cell(row, header_map, "Issue_Type")
                summary = self._get_cell(row, header_map, "Short_Description")
                detailed_description = self._get_cell(row, header_map, "Detailed_Description")
                status = self._get_cell(row, header_map, "Status")
                priority = self._get_cell(row, header_map, "Priority")
                created = self._get_cell(row, header_map, "Created_At")
                updated = self._get_cell(row, header_map, "Updated_At")
                resolution_notes = self._get_cell(row, header_map, "Resolution_Notes")
                root_cause = self._get_cell(row, header_map, "Root_Cause")
                workaround = self._get_cell(row, header_map, "Workaround")
                environment = self._get_cell(row, header_map, "Environment")
                category = self._get_cell(row, header_map, "Category")
                subcategory = self._get_cell(row, header_map, "Subcategory")
                config_item = self._get_cell(row, header_map, "Configuration_Item")
                assignment_group = self._get_cell(row, header_map, "Assignment_Group")
                assigned_to = self._get_cell(row, header_map, "Assigned_To")
                requester = self._get_cell(row, header_map, "Requester")
                channel = self._get_cell(row, header_map, "Channel")

                description_block = f"""
Detailed Description: {detailed_description}
Resolution Notes: {resolution_notes}
Root Cause: {root_cause}
Workaround: {workaround}
Environment: {environment}
Category: {category}
Subcategory: {subcategory}
Configuration Item: {config_item}
Assignment Group: {assignment_group}
Assigned To: {assigned_to}
Requester: {requester}
Channel: {channel}
""".strip()

                ticket_text = f"""
Issue Key: {incident_number}
Type: {issue_type}
Summary: {summary}
Status: {status}
Priority: {priority}
Created: {created}
Updated: {updated}

Description:
{description_block}
""".strip()

                docs.append(ticket_text)

        return docs

    def _get_cell(self, row, header_map, column_name: str) -> str:
        idx = header_map.get(column_name)
        if idx is None or idx >= len(row):
            return ""
        value = row[idx]
        return "" if value is None else str(value).strip()

    # ---------------- TEXT FIELD EXTRACTION ---------------- #

    def _extract_fields(self, text: str) -> Dict[str, str]:
        return {
            "issue_key": self._find(r"Issue\s*Key[:\-]\s*(.+)", text),
            "issue_type": self._find(r"Type[:\-]\s*(.+)", text),
            "summary": self._find(r"Summary[:\-]\s*(.+)", text),
            "status": self._find(r"Status[:\-]\s*(.+)", text),
            "priority": self._find(r"Priority[:\-]\s*(.+)", text),
            "created": self._find(r"Created[:\-]\s*(.+)", text),
            "updated": self._find(r"Updated[:\-]\s*(.+)", text),
        }

    def _find(self, pattern: str, text: str) -> str:
        match = re.search(pattern, text, re.IGNORECASE)
        return match.group(1).strip() if match else ""