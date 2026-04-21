"""
services/scheduler/ticket_scheduler.py

Background scheduler that polls both data sources for new tickets
and triggers auto-closure + resolution notification for each.

Flow
----
  New ticket detected
       │
       ├── RAGService.ask(description)    ← search knowledge base
       ├── Build HTML resolution table
       ├── Send to assignee email (or log if no SMTP)
       └── Evaluate auto-closure (confidence >= threshold)

Schedule (configurable in .env)
  SCHEDULER_JIRA_INTERVAL_MINUTES=1
  SCHEDULER_SHAREPOINT_INTERVAL_MINUTES=1
  SCHEDULER_ENABLED=true
"""

from dotenv import load_dotenv
load_dotenv()

import csv
import hashlib
import logging
import os
import sqlite3
import time
import traceback
from datetime import datetime, timedelta, timezone
from pathlib import Path
from app.core.db_path import get_db_path
from typing import Any, Dict, List, Optional

import requests
import urllib3
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.interval import IntervalTrigger
from requests.auth import HTTPBasicAuth

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)


# ── DB helpers ────────────────────────────────────────────────────────────────

def _get_db_path() -> str:
    # __file__ = backend/app/services/scheduler/ticket_scheduler.py
    # parents[0] = scheduler/  parents[1] = services/
    # parents[2] = app/        parents[3] = backend/
    base = Path(__file__).resolve().parents[3]
    db_dir = base / "database"
    db_dir.mkdir(parents=True, exist_ok=True)
    return str(db_dir / "ticket.db")


def _ensure_scheduler_table():
    conn = sqlite3.connect(_get_db_path())
    conn.execute("""
        CREATE TABLE IF NOT EXISTS scheduler_processed (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            tenant_id    TEXT NOT NULL,
            source_type  TEXT NOT NULL,
            ticket_id    TEXT NOT NULL,
            content_hash TEXT,
            processed_at TEXT NOT NULL,
            UNIQUE(tenant_id, source_type, ticket_id)
        )
    """)
    conn.commit()
    conn.close()


def _is_processed(tenant_id: str, source_type: str, ticket_id: str) -> bool:
    try:
        conn = sqlite3.connect(_get_db_path())
        cur = conn.execute(
            "SELECT 1 FROM scheduler_processed WHERE tenant_id=? AND source_type=? AND ticket_id=? LIMIT 1",
            (tenant_id, source_type, ticket_id),
        )
        found = cur.fetchone() is not None
        conn.close()
        return found
    except Exception as e:
        print(f"[Scheduler] DB check error: {e}")
        return False


def _mark_processed(tenant_id: str, source_type: str, ticket_id: str, content_hash: str = ""):
    try:
        conn = sqlite3.connect(_get_db_path())
        conn.execute(
            """
            INSERT OR IGNORE INTO scheduler_processed
                (tenant_id, source_type, ticket_id, content_hash, processed_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (tenant_id, source_type, ticket_id, content_hash,
             datetime.now(timezone.utc).isoformat()),
        )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[Scheduler] DB mark error: {e}")


def _content_hash(text: str) -> str:
    return hashlib.md5(text.encode("utf-8", errors="ignore")).hexdigest()


# ── Tenant config helpers ─────────────────────────────────────────────────────

def _get_all_tenant_ids() -> List[str]:
    base = Path(__file__).resolve().parents[3] / "config_store"
    if not base.exists():
        print(f"[Scheduler] config_store not found at {base}")
        return []
    ids = [f.stem.replace("_rag_config", "") for f in base.glob("*_rag_config.json")]
    print(f"[Scheduler] Tenants found: {ids}")
    return ids


def _get_raw_config(tenant_id: str) -> Optional[Dict]:
    """
    Read the raw rag_config.json directly — NOT via AIConfigRepository
    because that strips out data_sources. We need data_sources here
    to know which Jira / SharePoint sources to poll.
    """
    import json
    config_path = Path(__file__).resolve().parents[3] / "config_store" / f"{tenant_id}_rag_config.json"
    print(f"[Scheduler] Reading config from: {config_path}")
    print(f"[Scheduler] Config file exists: {config_path.exists()}")

    if not config_path.exists():
        print(f"[Scheduler] ❌ Config file not found for tenant {tenant_id}")
        return None

    with open(config_path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    sources = raw.get("data_sources", [])
    print(f"[Scheduler] data_sources count: {len(sources)}")
    for s in sources:
        print(f"[Scheduler]   -> source_type={s.get('source_type')!r}  is_enabled={s.get('is_enabled')!r}")

    return raw


# ── Core processor ────────────────────────────────────────────────────────────

def _process_new_ticket(
    tenant_id: str,
    source_type: str,
    ticket_id: str,
    description: str,
    assignee_email: Optional[str],
):
    """
    For a newly detected ticket — delegates to the LangGraph agent
    which autonomously decides: search → email → close or escalate.
    Falls back to the legacy pipeline if the agent is unavailable.
    """
    print(f"\n{'='*60}")
    print(f"[Scheduler] 🎫 NEW TICKET DETECTED")
    print(f"[Scheduler]    Ticket ID   : {ticket_id}")
    print(f"[Scheduler]    Source      : {source_type}")
    print(f"[Scheduler]    Tenant      : {tenant_id}")
    print(f"[Scheduler]    Assignee    : {assignee_email or 'Not assigned'}")
    print(f"[Scheduler]    Description : {description[:120]}...")
    print(f"{'='*60}")

    try:
        # ── 🤖 Multi-Agent System ─────────────────────────────────────────
        from app.services.agent.agents.coordinator_agent import run_multi_agent_system

        result = run_multi_agent_system(
            tenant_id      = tenant_id,
            ticket_id      = ticket_id,
            source_type    = source_type,
            description    = description,
            assignee_email = assignee_email,
        )

        print(f"[Scheduler] 🤖 Multi-agent system completed:")
        print(f"[Scheduler]    Ingestion   : {result['ingestion_status']}")
        print(f"[Scheduler]    Resolution  : {result['resolution_status']} "
              f"(conf={result['best_confidence']:.4f})")
        print(f"[Scheduler]    Notification: {result['notification_status']}")
        print(f"[Scheduler]    Closure     : {result['closure_decision']}")
        print(f"[Scheduler]    Summary     : {result['final_summary']}")
        if result.get('errors'):
            print(f"[Scheduler]    Errors      : {result['errors']}")

    except ImportError:
        # ── 🔁 Fallback: legacy hardcoded pipeline ────────────────────────
        print(f"[Scheduler] ⚠️  Multi-agent not available — using legacy pipeline")
        _process_new_ticket_legacy(
            tenant_id, source_type, ticket_id, description, assignee_email
        )
    except Exception as exc:
        print(f"[Scheduler] ❌ Multi-agent failed for {ticket_id}: {exc}")
        traceback.print_exc()
        print(f"[Scheduler] 🔁 Falling back to legacy pipeline...")
        _process_new_ticket_legacy(
            tenant_id, source_type, ticket_id, description, assignee_email
        )

    # ── Mark processed regardless of path ────────────────────────────────
    _mark_processed(tenant_id, source_type, ticket_id)
    print(f"[Scheduler] ✅ Ticket {ticket_id} fully processed and marked as done")
    print(f"{'='*60}\n")


def _process_new_ticket_legacy(
    tenant_id: str,
    source_type: str,
    ticket_id: str,
    description: str,
    assignee_email: Optional[str],
):
    """
    Legacy hardcoded pipeline — used as fallback when LangGraph is unavailable.
    """
    # ── Step 1: Search RAG vector DB ─────────────────────────────────────────
    print(f"[Scheduler] 🔍 Step 1: Searching vector DB for matching resolutions...")
    try:
        from app.repositories.rag_admin_repository import RagAdminRepository
        from app.services.ingestion_service import IngestionService

        repo = RagAdminRepository()
        ingestion_service = IngestionService(repo)
        retrieval_result = ingestion_service.query(tenant_id, description)
        tickets = retrieval_result.get("tickets", [])
        print(f"[Scheduler] ✅ RAG search complete — {len(tickets)} matching resolution(s) found")

        if not tickets:
            print(f"[Scheduler] ⚠️  No matching resolutions found in knowledge base for ticket {ticket_id}")
        else:
            for i, t in enumerate(tickets[:3], 1):
                print(f"[Scheduler]    Match {i}: {t.get('ticket_id','')} | "
                      f"confidence={t.get('confidence_score', 0):.4f} | "
                      f"resolution={str(t.get('resolution',''))[:60]}")

    except Exception as exc:
        print(f"[Scheduler] ❌ RAG search failed: {exc}")
        traceback.print_exc()
        tickets = []

    # ── Step 2 & 3: Build table and send notification ─────────────────────────
    print(f"[Scheduler] 📧 Step 2: Building resolution table and sending to assignee...")
    try:
        from app.schemas.notification import NotifyRequest
        from app.services.notification.notification_service import NotificationService

        result = NotificationService().notify_on_ticket_created(
            NotifyRequest(
                tenant_id=tenant_id,
                ticket_id=ticket_id,
                source_type=source_type,
                description=description,
                assignee_email=assignee_email,
                top_k=5,
                prefetched_tickets=tickets,
            )
        )
        print(f"[Scheduler] ✅ Notification sent!")
        print(f"[Scheduler]    Channel     : {result.channel}")
        print(f"[Scheduler]    Status      : {result.status}")
        print(f"[Scheduler]    Resolutions : {len(result.resolutions)} row(s) in table")
        print(f"[Scheduler]    Message     : {result.message}")

    except Exception as exc:
        print(f"[Scheduler] ❌ Notification failed for {ticket_id}: {exc}")
        traceback.print_exc()

    # ── Step 4: Auto-closure evaluation ──────────────────────────────────────
    print(f"[Scheduler] 🔒 Step 3: Evaluating auto-closure...")
    try:
        from app.schemas.ticket_lifecycle import AutoCloseRequest
        from app.services.ticket_lifecycle.auto_closure_service import (
            DEFAULT_CONFIDENCE_THRESHOLD,
            AutoClosureService,
        )

        closure = AutoClosureService().evaluate(
            AutoCloseRequest(
                tenant_id=tenant_id,
                ticket_id=ticket_id,
                source_type=source_type,
                description=description,
                assignee_email=assignee_email,
                confidence_threshold=DEFAULT_CONFIDENCE_THRESHOLD,
            )
        )
        print(f"[Scheduler]    Auto-closed : {closure.auto_closed}")
        print(f"[Scheduler]    Confidence  : {closure.confidence_score:.4f}")
        print(f"[Scheduler]    Matched     : {closure.matched_ticket_id or 'None'}")
        print(f"[Scheduler]    Reason      : {closure.reason}")

    except Exception as exc:
        print(f"[Scheduler] ❌ Auto-closure failed for {ticket_id}: {exc}")
        traceback.print_exc()


# ── Jira Poller ───────────────────────────────────────────────────────────────

class JiraPoller:
    """
    Polls Jira for tickets created within the last N minutes using
    Jira Cloud REST API v3 with JQL relative date format.
    """

    def poll(self, tenant_id: str, source: Dict[str, Any], interval_minutes: int) -> List[Dict]:
        base_url    = (source.get("source_url") or "").strip().rstrip("/")
        username    = (source.get("username") or "").strip()
        token       = (source.get("token") or "").strip()
        project_key = (source.get("project_key") or "").strip()

        print(f"[Jira Poller] tenant={tenant_id} project={project_key} url={base_url}")

        if not base_url or not username or not token:
            print(f"[Jira Poller] ❌ Missing credentials for tenant {tenant_id}")
            return []

        # Jira Cloud JQL relative time format: created >= "-15m"
        # Lookback window = 600 minutes (10 hours)
        # scheduler_processed table prevents duplicate processing
        jql = f'project = "{project_key}" AND created >= "-600m" ORDER BY created DESC'
        print(f"[Jira Poller] JQL: {jql}")

        search_url = f"{base_url}/rest/api/3/search/jql"
        auth = HTTPBasicAuth(username, token)

        try:
            resp = requests.post(
                search_url,
                auth=auth,
                headers={"Accept": "application/json", "Content-Type": "application/json"},
                json={
                    "jql": jql,
                    "maxResults": 50,
                    "fields": [
                        "summary", "description", "status",
                        "priority", "issuetype", "created", "assignee",
                    ],
                },
                timeout=30,
                verify=False,
            )

            print(f"[Jira Poller] Response status: {resp.status_code}")

            if resp.status_code == 401:
                print(f"[Jira Poller] ❌ Authentication failed — check username/token")
                return []
            if resp.status_code == 400:
                print(f"[Jira Poller] ❌ Bad request — JQL error: {resp.text[:300]}")
                return []
            if resp.status_code != 200:
                print(f"[Jira Poller] ❌ HTTP {resp.status_code}: {resp.text[:200]}")
                return []

            data   = resp.json()
            issues = data.get("issues", [])
            print(f"[Jira Poller] Total issues returned by Jira: {len(issues)}")

        except Exception as exc:
            print(f"[Jira Poller] ❌ Request failed: {exc}")
            traceback.print_exc()
            return []

        new_tickets = []
        for issue in issues:
            ticket_id = issue.get("key", "")
            if not ticket_id:
                continue

            if _is_processed(tenant_id, "jira", ticket_id):
                print(f"[Jira Poller] ⏭ Already processed: {ticket_id}")
                continue

            fields        = issue.get("fields", {})
            summary       = (fields.get("summary") or "").strip()
            raw_desc      = fields.get("description")
            description   = self._extract_description(raw_desc) or summary
            assignee      = fields.get("assignee") or {}
            assignee_email = assignee.get("emailAddress")

            print(f"[Jira Poller] 🆕 New ticket: {ticket_id} | "
                  f"desc={description[:80]} | assignee={assignee_email}")

            if not description:
                print(f"[Jira Poller] ⚠ Skipping {ticket_id} — no description or summary")
                continue

            new_tickets.append({
                "ticket_id":     ticket_id,
                "description":   description,
                "assignee_email": assignee_email,
            })

        print(f"[Jira Poller] New unprocessed tickets: {len(new_tickets)}")
        return new_tickets

    def _extract_description(self, node: Any) -> str:
        """Recursively extract plain text from Jira ADF format."""
        if isinstance(node, str):
            return node.strip()
        if not isinstance(node, dict):
            return ""
        texts = []

        def _walk(n):
            if isinstance(n, dict):
                if n.get("type") == "text":
                    texts.append(n.get("text", ""))
                for child in n.get("content", []):
                    _walk(child)
            elif isinstance(n, list):
                for item in n:
                    _walk(item)

        _walk(node)
        return " ".join(t for t in texts if t).strip()


# ── SharePoint Local Poller ───────────────────────────────────────────────────

class SharePointLocalPoller:
    """
    Scans SharePoint local folder for new ticket rows.
    Supports: .xlsx, .csv, .txt, .md, .pdf, .docx
    Uses content hash per row/file to detect new entries.
    """

    SUPPORTED = {".xlsx", ".csv", ".txt", ".md", ".pdf", ".docx"}

    # Tracks last modified time per file path to avoid re-scanning unchanged files
    _file_mod_cache: Dict[str, float] = {}

    def poll(self, tenant_id: str, source: Dict[str, Any]) -> List[Dict]:
        folder_path = (source.get("source_url") or source.get("folder_path") or "").strip()

        if not folder_path:
            return []

        base = Path(folder_path)
        if not base.exists():
            print(f"[SP Poller] ❌ Folder not found: {folder_path}")
            return []

        new_tickets = []

        for file_path in base.rglob("*"):
            if not file_path.is_file():
                continue
            if file_path.suffix.lower() not in self.SUPPORTED:
                continue

            # ── Check file modification time ──────────────────────────────
            # Only scan if file was modified since last poll
            file_key = str(file_path)
            current_mtime = file_path.stat().st_mtime
            last_mtime = self._file_mod_cache.get(file_key, 0)

            if current_mtime <= last_mtime:
                # File not modified since last scan — skip silently
                continue

            print(f"[SP Poller] 📄 File modified — scanning: {file_path.name}")
            # Update cache before scanning
            self._file_mod_cache[file_key] = current_mtime

            try:
                if file_path.suffix.lower() == ".xlsx":
                    found = self._scan_xlsx(tenant_id, file_path)
                elif file_path.suffix.lower() == ".csv":
                    found = self._scan_csv(tenant_id, file_path)
                else:
                    found = self._scan_text_file(tenant_id, file_path)
                    found = [found] if found else []

                if found:
                    print(f"[SP Poller] {file_path.name} → {len(found)} new ticket(s)")
                new_tickets.extend(found)

            except Exception as exc:
                print(f"[SP Poller] ❌ Error scanning {file_path.name}: {exc}")
                traceback.print_exc()

        if new_tickets:
            print(f"[SP Poller] Total new tickets found: {len(new_tickets)}")
        return new_tickets

    def _scan_xlsx(self, tenant_id: str, file_path: Path) -> List[Dict]:
        from openpyxl import load_workbook

        wb = load_workbook(filename=str(file_path), data_only=True, read_only=True)
        results = []

        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            headers    = [str(h).strip() if h else "" for h in rows[0]]
            header_map = {h: i for i, h in enumerate(headers) if h}

            if "Incident_Number" not in header_map:
                print(f"[SP Poller] Sheet '{ws.title}' skipped — no Incident_Number column")
                continue

            for row in rows[1:]:
                incident_num = self._cell(row, header_map, "Incident_Number")
                if not incident_num:
                    continue

                ticket_id = f"SP-{incident_num}"
                row_text  = " | ".join(str(v) for v in row if v)
                chash     = _content_hash(row_text)

                if _is_processed(tenant_id, "sharepoint_local", ticket_id):
                    continue

                description = (
                    self._cell(row, header_map, "Detailed_Description")
                    or self._cell(row, header_map, "Short_Description")
                    or row_text[:300]
                )
                assignee_email = self._cell(row, header_map, "Assignee_Email") or None

                print(f"[SP Poller] 🆕 New XLSX row: {ticket_id} | desc={description[:80]}")
                _mark_processed(tenant_id, "sharepoint_local", ticket_id, chash)
                results.append({
                    "ticket_id":     ticket_id,
                    "description":   description,
                    "assignee_email": assignee_email,
                    "content_hash":  chash,
                })

        wb.close()
        return results

    def _scan_csv(self, tenant_id: str, file_path: Path) -> List[Dict]:
        results = []
        text    = file_path.read_text(encoding="utf-8", errors="ignore")
        reader  = csv.DictReader(text.splitlines())

        for i, row in enumerate(reader):
            row_text  = " | ".join(f"{k}:{v}" for k, v in row.items() if v)
            chash     = _content_hash(row_text)
            ticket_id = (
                row.get("Incident_Number")
                or row.get("ticket_id")
                or row.get("id")
                or f"{file_path.stem}-row-{i+1}"
            )
            ticket_id = f"SP-{ticket_id}"

            if _is_processed(tenant_id, "sharepoint_local", ticket_id):
                continue

            description = (
                row.get("Detailed_Description")
                or row.get("Short_Description")
                or row.get("description")
                or row_text[:300]
            )
            assignee_email = row.get("Assignee_Email") or row.get("assignee_email") or None

            print(f"[SP Poller] 🆕 New CSV row: {ticket_id} | desc={description[:80]}")
            _mark_processed(tenant_id, "sharepoint_local", ticket_id, chash)
            results.append({
                "ticket_id":     ticket_id,
                "description":   description,
                "assignee_email": assignee_email,
                "content_hash":  chash,
            })

        return results

    def _scan_text_file(self, tenant_id: str, file_path: Path) -> Optional[Dict]:
        suffix = file_path.suffix.lower()

        if suffix == ".pdf":
            from pypdf import PdfReader
            reader  = PdfReader(str(file_path))
            content = "\n".join(p.extract_text() or "" for p in reader.pages).strip()
        elif suffix == ".docx":
            from docx import Document
            doc     = Document(str(file_path))
            content = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        else:
            content = file_path.read_text(encoding="utf-8", errors="ignore").strip()

        if not content:
            return None

        chash     = _content_hash(content)
        ticket_id = f"SP-FILE-{file_path.stem}"

        if _is_processed(tenant_id, "sharepoint_local", ticket_id):
            return None

        print(f"[SP Poller] 🆕 New file: {ticket_id} | content={content[:80]}")
        _mark_processed(tenant_id, "sharepoint_local", ticket_id, chash)
        return {
            "ticket_id":     ticket_id,
            "description":   content[:500],
            "assignee_email": None,
            "content_hash":  chash,
        }

    def _cell(self, row, header_map: Dict, col: str) -> str:
        idx = header_map.get(col)
        if idx is None or idx >= len(row):
            return ""
        return "" if row[idx] is None else str(row[idx]).strip()


# ── Scheduler job functions ───────────────────────────────────────────────────

def run_jira_scheduler_job():
    """APScheduler job — polls Jira for all enabled tenants."""
    global _jira_poller
    if _jira_poller is None:
        _jira_poller = JiraPoller()
    poller = _jira_poller

    interval = int(os.getenv("SCHEDULER_JIRA_INTERVAL_MINUTES", "15"))

    for tenant_id in _get_all_tenant_ids():
        raw = _get_raw_config(tenant_id)
        if not raw:
            print(f"[Scheduler] No config for tenant {tenant_id}")
            continue

        sources = raw.get("data_sources", [])
        jira_sources = [
            s for s in sources
            if (s.get("source_type") or "").lower() == "jira"
            and s.get("is_enabled", True)
        ]
        print(f"[Scheduler] Tenant {tenant_id} — Total data sources: {len(sources)}")
        for s in sources:
            print(f"[Scheduler]   source_type={s.get('source_type')!r} is_enabled={s.get('is_enabled')!r}")
        print(f"[Scheduler] Tenant {tenant_id} — Jira sources matched: {len(jira_sources)}")

        for source in jira_sources:
            new_tickets = poller.poll(tenant_id, source, interval)
            if not new_tickets:
                print(f"[Scheduler] No new Jira tickets for tenant {tenant_id} — skipping.")
                continue
            for t in new_tickets:
                _process_new_ticket(
                    tenant_id=tenant_id,
                    source_type="jira",
                    ticket_id=t["ticket_id"],
                    description=t["description"],
                    assignee_email=t.get("assignee_email"),
                )
                time.sleep(1)

    print("[Scheduler] ✅ Jira polling job completed")
    logger.info("[Scheduler] ✅ Jira polling job completed")


# Module-level singleton pollers so _file_mod_cache persists across job runs
_jira_poller = None
_sp_poller = None

def run_sharepoint_scheduler_job():
    """APScheduler job — scans SharePoint local folders for all enabled tenants."""
    global _sp_poller
    if _sp_poller is None:
        _sp_poller = SharePointLocalPoller()
    poller = _sp_poller
    found_any = False


    for tenant_id in _get_all_tenant_ids():
        raw = _get_raw_config(tenant_id)
        if not raw:
            continue

        sources = raw.get("data_sources", [])
        sp_sources = [
            s for s in sources
            if (s.get("source_type") or "").lower() == "sharepoint_local"
            and s.get("is_enabled", True)
        ]
        for source in sp_sources:
            new_tickets = poller.poll(tenant_id, source)
            if not new_tickets:
                continue
            found_any = True
            for t in new_tickets:
                _process_new_ticket(
                    tenant_id=tenant_id,
                    source_type="sharepoint_local",
                    ticket_id=t["ticket_id"],
                    description=t["description"],
                    assignee_email=t.get("assignee_email"),
                )
                time.sleep(1)

    if found_any:
        print("[Scheduler] ✅ SharePoint Local polling job completed")
        logger.info("[Scheduler] ✅ SharePoint Local polling job completed")


# ── Scheduler lifecycle ───────────────────────────────────────────────────────

_scheduler: Optional[BackgroundScheduler] = None


def start_scheduler():
    """Start background scheduler. Called from main.py on startup."""
    global _scheduler

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    try:
        enabled = os.getenv("SCHEDULER_ENABLED", "true").lower()
        print(f"[Scheduler] SCHEDULER_ENABLED={enabled}")

        if enabled != "true":
            print("[Scheduler] Disabled — skipping.")
            return

        _ensure_scheduler_table()
        print("[Scheduler] scheduler_processed table ready.")

        jira_interval = int(os.getenv("SCHEDULER_JIRA_INTERVAL_MINUTES", "15"))
        sp_interval   = int(os.getenv("SCHEDULER_SHAREPOINT_INTERVAL_MINUTES", "15"))
        print(f"[Scheduler] Jira interval={jira_interval}m  SharePoint interval={sp_interval}m")

        _scheduler = BackgroundScheduler(timezone="UTC")

        _scheduler.add_job(
            run_jira_scheduler_job,
            trigger=IntervalTrigger(minutes=jira_interval),
            id="jira_poller",
            name="Jira New Ticket Poller",
            replace_existing=True,
            next_run_time=datetime.now(timezone.utc),
        )

        _scheduler.add_job(
            run_sharepoint_scheduler_job,
            trigger=IntervalTrigger(minutes=sp_interval),
            id="sharepoint_local_poller",
            name="SharePoint Local New Ticket Poller",
            replace_existing=True,
            next_run_time=datetime.now(timezone.utc),
        )

        _scheduler.start()
        print(f"[Scheduler] ✅ Started — Jira every {jira_interval}m, SharePoint every {sp_interval}m")

    except Exception as exc:
        print(f"[Scheduler] ❌ Failed to start: {exc}")
        traceback.print_exc()


def stop_scheduler():
    """Gracefully stop the scheduler on app shutdown."""
    global _scheduler
    if _scheduler and _scheduler.running:
        _scheduler.shutdown(wait=False)
        print("[Scheduler] Stopped.")